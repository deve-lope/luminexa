"""Tests for business data export functionality."""

from decimal import Decimal
from io import BytesIO, StringIO
import csv
import json
import zipfile

from django.contrib.auth import get_user_model
from django.test import TestCase, override_settings
from django.utils import timezone
from rest_framework import status
from rest_framework.test import APIClient

from businesses.models import Organization, OrganizationLocation, OrganizationMembership
from jobs.models import (
    Booking,
    Invoice,
    JobCostLine,
    Service,
    ServiceCategory,
    ServiceRequestMessage,
    Task,
)

User = get_user_model()


@override_settings(SECURE_SSL_REDIRECT=False)
class OrganizationDataExportTestCase(TestCase):
    """Test data export permissions, formats, and content."""

    def setUp(self):
        self.client = APIClient()

        self.owner = User.objects.create_user(
            email='owner@test.local',
            full_name='Owner User',
            phone='+15551234567',
        )
        self.staff = User.objects.create_user(
            email='staff@test.local',
            full_name='Staff User',
        )
        self.customer = User.objects.create_user(
            email='customer@test.local',
            full_name='Customer User',
            phone='+15559876543',
            default_service_address='123 Main St, Toronto, ON M1A 1A1',
        )
        self.other_user = User.objects.create_user(
            email='other@test.local',
            full_name='Other User',
        )

        self.org = Organization.objects.create(
            name='Test Service Co',
            slug='test-service',
            tagline='Test tagline',
            description='Test description',
            timezone='America/Toronto',
            booking_policy=Organization.BookingPolicy.INSTANT,
            cancel_cutoff_hours=24,
            concurrent_capacity=2,
            scheduling_mode=Organization.SchedulingMode.FLEXI,
            service_address='456 Business Ave',
            service_city='Toronto',
            service_state='ON',
            service_postal_code='M2B 2B2',
            service_latitude=Decimal('43.700000'),
            service_longitude=Decimal('-79.400000'),
            service_radius_miles=Decimal('25.0'),
            subscription_plan='pro_monthly',
            subscription_status='active',
        )

        OrganizationMembership.objects.create(
            user=self.owner,
            organization=self.org,
            role=OrganizationMembership.Role.OWNER,
        )
        OrganizationMembership.objects.create(
            user=self.staff,
            organization=self.org,
            role=OrganizationMembership.Role.STAFF,
        )
        OrganizationMembership.objects.create(
            user=self.customer,
            organization=self.org,
            role=OrganizationMembership.Role.CUSTOMER,
            customer_status=OrganizationMembership.CustomerStatus.APPROVED,
            provider_notes='VIP customer',
        )

        OrganizationLocation.objects.create(
            organization=self.org,
            address='789 Branch St',
            city='Mississauga',
            state='ON',
            postal_code='L5A 1A1',
            latitude=Decimal('43.600000'),
            longitude=Decimal('-79.600000'),
            radius_miles=Decimal('15.0'),
            is_primary=False,
        )

        self.category = ServiceCategory.objects.create(
            organization=self.org,
            name='Cleaning',
            sort_order=1,
        )
        self.service = Service.objects.create(
            organization=self.org,
            category=self.category,
            name='House Cleaning',
            description='Full house cleaning service',
            duration_minutes=120,
            pricing_type=Service.PricingType.FIXED,
            base_price=Decimal('100.00'),
        )

        now = timezone.now()
        self.booking = Booking.objects.create(
            organization=self.org,
            customer=self.customer,
            service=self.service,
            start_at=now,
            end_at=now + timezone.timedelta(hours=2),
            status=Booking.Status.COMPLETED,
            service_address='123 Main St, Toronto, ON M1A 1A1',
            quote_amount=Decimal('100.00'),
            customer_notes='Please ring doorbell',
        )

        Invoice.objects.create(
            booking=self.booking,
            number='INV-001',
            description='House Cleaning',
            amount=Decimal('100.00'),
            subtotal=Decimal('88.50'),
            tax_total=Decimal('11.50'),
            currency='CAD',
            status=Invoice.Status.PAID,
            paid_at=now,
        )

        JobCostLine.objects.create(
            booking=self.booking,
            kind=JobCostLine.Kind.MATERIAL,
            description='Cleaning supplies',
            quantity=Decimal('1.0'),
            unit_cost=Decimal('25.00'),
        )

        ServiceRequestMessage.objects.create(
            booking=self.booking,
            sender=self.customer,
            body='When will you arrive?',
        )

        Task.objects.create(
            organization=self.org,
            job=self.booking,
            title='Follow up on job',
            notes='Check customer satisfaction',
        )

        self.export_url = f'/api/v1/organizations/{self.org.slug}/data-export/'

    def _get(self, user, **params):
        self.client.force_authenticate(user=user)
        # DRF reserves ?format= for content negotiation; use export_format.
        if 'format' in params:
            params['export_format'] = params.pop('format')
        return self.client.get(self.export_url, params, HTTP_HOST='localhost')

    def test_non_owner_denied(self):
        for user in (self.staff, self.customer, self.other_user):
            response = self._get(user, format='json')
            self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)
            self.assertIn('owner', str(response.data).lower())

    def test_free_tier_denied(self):
        self.org.subscription_plan = 'free'
        self.org.save(update_fields=['subscription_plan'])

        response = self._get(self.owner, format='json')
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)
        self.assertIn('pro', str(response.data).lower())

    def test_json_export_structure(self):
        response = self._get(self.owner, format='json')

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response['Content-Type'], 'application/json')
        self.assertIn('attachment', response['Content-Disposition'])
        self.assertIn(f'{self.org.slug}-export', response['Content-Disposition'])
        self.assertIn('.json', response['Content-Disposition'])

        data = json.loads(response.content)

        for key in (
            'export_date', 'export_version', 'organization', 'locations', 'staff',
            'categories', 'services', 'schedule', 'customers', 'bookings',
            'invoices', 'job_costs', 'tasks',
        ):
            self.assertIn(key, data)

        self.assertEqual(data['organization']['name'], 'Test Service Co')
        self.assertEqual(data['organization']['slug'], 'test-service')
        self.assertEqual(data['organization']['concurrent_capacity'], 2)

        self.assertEqual(len(data['locations']), 1)
        self.assertEqual(data['locations'][0]['city'], 'Mississauga')

        staff_emails = [s['email'] for s in data['staff']]
        self.assertIn('owner@test.local', staff_emails)
        self.assertIn('staff@test.local', staff_emails)

        self.assertEqual(len(data['customers']), 1)
        cust = data['customers'][0]
        self.assertEqual(cust['email'], 'customer@test.local')
        self.assertEqual(cust['provider_notes'], 'VIP customer')
        self.assertEqual(cust['total_bookings'], 1)

        self.assertEqual(len(data['services']), 1)
        self.assertEqual(data['services'][0]['name'], 'House Cleaning')
        self.assertEqual(data['services'][0]['base_price'], '100.00')

        self.assertEqual(len(data['bookings']), 1)
        bk = data['bookings'][0]
        self.assertEqual(bk['customer_email'], 'customer@test.local')
        self.assertEqual(bk['service_name'], 'House Cleaning')
        self.assertEqual(bk['status'], Booking.Status.COMPLETED)
        self.assertEqual(bk['customer_notes'], 'Please ring doorbell')
        self.assertEqual(len(bk['messages']), 1)
        self.assertEqual(bk['messages'][0]['body'], 'When will you arrive?')

        self.assertEqual(len(data['invoices']), 1)
        self.assertEqual(data['invoices'][0]['number'], 'INV-001')
        self.assertEqual(data['invoices'][0]['amount'], '100.00')

        self.assertEqual(len(data['job_costs']), 1)
        self.assertEqual(data['job_costs'][0]['total_cost'], '25.00')

        self.assertEqual(len(data['tasks']), 1)
        self.assertEqual(data['tasks'][0]['title'], 'Follow up on job')

    def test_csv_export_format(self):
        response = self._get(self.owner, format='csv')

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response['Content-Type'], 'application/zip')
        self.assertIn('.zip', response['Content-Disposition'])

        with zipfile.ZipFile(BytesIO(response.content), 'r') as zf:
            names = zf.namelist()
            for expected in (
                'organization.csv', 'customers.csv', 'bookings.csv',
                'invoices.csv', 'services.csv',
            ):
                self.assertIn(expected, names)

            customers_csv = zf.read('customers.csv').decode('utf-8')
            rows = list(csv.DictReader(StringIO(customers_csv)))
            self.assertEqual(len(rows), 1)
            self.assertEqual(rows[0]['email'], 'customer@test.local')

    def test_excel_export_format(self):
        response = self._get(self.owner, format='excel')

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        self.assertIn('.xlsx', response['Content-Disposition'])

        from openpyxl import load_workbook

        wb = load_workbook(BytesIO(response.content))
        for sheet in ('Organization', 'Customers', 'Bookings', 'Invoices', 'Services'):
            self.assertIn(sheet, wb.sheetnames)

        customers_sheet = wb['Customers']
        headers = [cell.value for cell in customers_sheet[1]]
        self.assertIn('email', headers)
        email_col = headers.index('email') + 1
        self.assertEqual(customers_sheet.cell(2, email_col).value, 'customer@test.local')

    def test_empty_org_exports_cleanly(self):
        minimal_org = Organization.objects.create(
            name='Empty Org',
            slug='empty-org',
            subscription_plan='pro_monthly',
            subscription_status='active',
        )
        minimal_owner = User.objects.create_user(
            email='minimal@test.local',
            full_name='Minimal User',
        )
        OrganizationMembership.objects.create(
            user=minimal_owner,
            organization=minimal_org,
            role=OrganizationMembership.Role.OWNER,
        )

        self.client.force_authenticate(user=minimal_owner)
        response = self.client.get(
            f'/api/v1/organizations/{minimal_org.slug}/data-export/',
            {'export_format': 'json'},
            HTTP_HOST='localhost',
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        data = json.loads(response.content)
        self.assertEqual(data['organization']['name'], 'Empty Org')
        self.assertEqual(len(data['bookings']), 0)
        self.assertEqual(len(data['customers']), 0)
        self.assertEqual(len(data['services']), 0)

    def test_customer_pii_included(self):
        response = self._get(self.owner, format='json')
        data = json.loads(response.content)
        cust = data['customers'][0]

        self.assertEqual(cust['email'], 'customer@test.local')
        self.assertEqual(cust['phone'], '+15559876543')
        self.assertEqual(cust['default_service_address'], '123 Main St, Toronto, ON M1A 1A1')

    def test_no_sensitive_credentials_in_export(self):
        self.org.stripe_customer_id = 'cus_test123'
        self.org.stripe_account_id = 'acct_test456'
        self.org.qbo_realm_id = 'qbo_realm_789'
        self.org.qbo_access_token = 'secret_access_token'
        self.org.qbo_refresh_token = 'secret_refresh_token'
        self.org.save()

        response = self._get(self.owner, format='json')
        data = json.loads(response.content)
        org_data = data['organization']
        org_blob = json.dumps(org_data)

        for key in (
            'stripe_customer_id', 'stripe_account_id', 'qbo_access_token',
            'qbo_refresh_token', 'qbo_realm_id',
        ):
            self.assertNotIn(key, org_data)

        self.assertNotIn('cus_test123', org_blob)
        self.assertNotIn('acct_test456', org_blob)
        self.assertNotIn('secret_access_token', org_blob)
        self.assertNotIn('secret_refresh_token', org_blob)

    def test_invalid_format_rejected(self):
        response = self._get(self.owner, format='pdf')
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_default_format_is_json(self):
        response = self._get(self.owner)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response['Content-Type'], 'application/json')
