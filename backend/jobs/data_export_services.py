"""
Organization data export for business migration and backup.

Collects complete operational data for a service provider organization
and exports in JSON, CSV (ZIP), or Excel formats.
"""

from __future__ import annotations

import csv
import io
import json
import zipfile
from datetime import timezone as dt_timezone
from typing import Any

from django.utils import timezone

from businesses.models import (
    Organization,
    OrganizationGalleryImage,
    OrganizationLocation,
    OrganizationMembership,
)

from .models import (
    AvailabilitySlot,
    Booking,
    BookingStatusEvent,
    CustomerServiceInquiry,
    Invoice,
    JobCostLine,
    Service,
    ServiceCategory,
    ServiceGalleryImage,
    ServiceRequestMessage,
    ServiceReview,
    Task,
    UnavailableBlock,
    WeeklyScheduleBlock,
)


MIGRATION_README = """\
Luminexa business data export — quick start
==========================================

This download is a full snapshot of your business data for backup or migration.
It is not a one-click import into another booking app. Use it to recreate your
customers, catalog, schedule, and history wherever you go next.

What you get
------------
- organization — business profile, booking settings, service area
- locations — branches / service pins and radii
- services / categories — your catalog (prices, duration, fulfillment)
- customers — people who booked with you (name, email, phone, address, notes)
- bookings — appointments, quotes, addresses, notes, status history
- messages — conversation threads on bookings and custom requests
- inquiries — custom service requests that were not catalog bookings
- invoices / job_costs — billing and your internal job costing
- schedule_hours / unavailable_blocks — weekly hours and time off
- reviews / tasks / staff — ratings, to-dos, and team roster
- gallery — image URLs (images stay on Luminexa hosting; save copies if needed)

Formats
-------
- JSON — one file; best for developers or technical imports
- CSV (ZIP) — open each .csv in Excel or Google Sheets
- Excel — one workbook with a sheet per data type (start with the README sheet)

Suggested migration steps
-------------------------
1. Keep a secure copy of this export (password-protect if you store it long-term).
2. Open customers and import contacts into your new CRM / phone / email tool.
3. Recreate services and weekly hours in your new booking system from the catalog
   and schedule files.
4. Use bookings and invoices as a historical archive (most tools cannot import
   Luminexa bookings 1:1).
5. After you confirm everything you need is here, cancel Luminexa from Billing
   if you are leaving.

Privacy
-------
Customer contact details are included because they booked with your business.
You are responsible for handling this data securely and in line with privacy laws.
Do not email the raw export unless the recipient is trusted.

Not included (by design)
------------------------
- Login codes and passwords
- Stripe / QuickBooks credentials
- Platform notification inbox items

Questions? Contact Luminexa support from your account settings.
"""


def migration_readme_text() -> str:
    return MIGRATION_README.strip() + '\n'


def collect_organization_data(organization: Organization) -> dict[str, Any]:
    """
    Gather all exportable data for an organization.
    
    Returns a dictionary with nested structures for all business entities.
    """
    export_date = timezone.now()
    
    # Organization profile
    org_data = {
        'id': organization.id,
        'name': organization.name,
        'slug': organization.slug,
        'public_ref': organization.public_ref,
        'tagline': organization.tagline,
        'description': organization.description,
        'timezone': organization.timezone or 'UTC',
        'booking_policy': organization.booking_policy,
        'cancel_cutoff_hours': organization.cancel_cutoff_hours,
        'concurrent_capacity': organization.concurrent_capacity,
        'scheduling_mode': organization.scheduling_mode,
        'schedule_valid_from': organization.schedule_valid_from.isoformat() if organization.schedule_valid_from else None,
        'schedule_valid_until': organization.schedule_valid_until.isoformat() if organization.schedule_valid_until else None,
        'service_address': organization.service_address,
        'service_city': organization.service_city,
        'service_state': organization.service_state,
        'service_postal_code': organization.service_postal_code,
        'service_latitude': float(organization.service_latitude) if organization.service_latitude else None,
        'service_longitude': float(organization.service_longitude) if organization.service_longitude else None,
        'service_radius_miles': float(organization.service_radius_miles) if organization.service_radius_miles else None,
        'logo_url': organization.logo.url if organization.logo else None,
        'banner_url': organization.banner.url if organization.banner else None,
        'business_types': [bt.name for bt in organization.business_types.all()],
        'created_at': organization.created_at.isoformat() if hasattr(organization, 'created_at') and organization.created_at else None,
    }
    
    # Locations / branches
    locations = []
    for loc in OrganizationLocation.objects.filter(organization=organization, is_active=True):
        locations.append({
            'id': loc.id,
            'address': loc.address,
            'city': loc.city,
            'state': loc.state,
            'postal_code': loc.postal_code,
            'latitude': float(loc.latitude) if loc.latitude else None,
            'longitude': float(loc.longitude) if loc.longitude else None,
            'radius_miles': float(loc.radius_miles) if loc.radius_miles else None,
            'is_primary': loc.is_primary,
        })
    
    # Gallery images
    gallery = []
    for img in OrganizationGalleryImage.objects.filter(organization=organization):
        gallery.append({
            'id': img.id,
            'image_url': img.image.url if img.image else None,
            'caption': img.caption,
            'sort_order': img.sort_order,
        })

    # Staff
    staff = []
    for mem in OrganizationMembership.objects.filter(
        organization=organization,
        role__in=[OrganizationMembership.Role.OWNER, OrganizationMembership.Role.STAFF],
    ).select_related('user'):
        staff.append({
            'id': mem.id,
            'user_id': mem.user_id,
            'full_name': mem.user.full_name,
            'email': mem.user.email,
            'phone': mem.user.phone,
            'role': mem.role,
        })

    # Service categories
    categories = []
    for cat in ServiceCategory.objects.filter(organization=organization):
        categories.append({
            'id': cat.id,
            'name': cat.name,
            'sort_order': cat.sort_order,
            'is_active': cat.is_active,
        })

    # Services
    services = []
    for svc in Service.objects.filter(organization=organization).select_related(
        'category',
    ).prefetch_related('gallery_images'):
        svc_gallery = []
        for img in svc.gallery_images.all():
            svc_gallery.append({
                'image_url': img.image.url if img.image else None,
                'sort_order': img.sort_order,
            })

        services.append({
            'id': svc.id,
            'category_id': svc.category_id,
            'category_name': svc.category.name if svc.category else None,
            'name': svc.name,
            'description': svc.description,
            'image_url': svc.image.url if svc.image else None,
            'duration_minutes': svc.duration_minutes,
            'pricing_type': svc.pricing_type,
            'base_price': str(svc.base_price),
            'price_max': str(svc.price_max) if svc.price_max else None,
            'show_price': svc.show_price,
            'quote_questions': svc.quote_questions,
            'allow_request': svc.allow_request,
            'fulfillment_kind': svc.fulfillment_kind,
            'is_active': svc.is_active,
            'sort_order': svc.sort_order,
            'gallery': svc_gallery,
            'created_at': svc.created_at.isoformat(),
        })

    # Schedule blocks
    weekly_hours = []
    for block in WeeklyScheduleBlock.objects.filter(organization=organization):
        weekly_hours.append({
            'id': block.id,
            'weekday': block.weekday,
            'start_time': str(block.start_time),
            'end_time': str(block.end_time),
            'is_active': block.is_active,
        })

    unavailable = []
    for block in UnavailableBlock.objects.filter(organization=organization):
        unavailable.append({
            'id': block.id,
            'start_at': block.start_at.isoformat(),
            'end_at': block.end_at.isoformat(),
            'note': block.note,
        })

    # Only export booked/pending slots (not all open generated slots)
    slots = []
    for slot in AvailabilitySlot.objects.filter(
        organization=organization,
        status__in=[AvailabilitySlot.Status.PENDING, AvailabilitySlot.Status.BOOKED],
    ).order_by('start_at'):
        slots.append({
            'id': slot.id,
            'start_at': slot.start_at.isoformat(),
            'end_at': slot.end_at.isoformat(),
            'status': slot.status,
            'capacity': slot.capacity,
        })

    # Customers (membership records with booking history)
    customers = []
    customer_memberships = OrganizationMembership.objects.filter(
        organization=organization,
        role=OrganizationMembership.Role.CUSTOMER,
    ).select_related('user')

    for mem in customer_memberships:
        booking_count = Booking.objects.filter(
            organization=organization,
            customer_id=mem.user_id,
        ).count()
        customers.append({
            'id': mem.id,
            'user_id': mem.user_id,
            'full_name': mem.user.full_name,
            'email': mem.user.email,
            'phone': mem.user.phone,
            'default_service_address': mem.user.default_service_address,
            'address_country': mem.user.address_country,
            'customer_status': mem.customer_status,
            'provider_notes': mem.provider_notes,
            'qbo_customer_id': mem.qbo_customer_id,
            'total_bookings': booking_count,
            'created_at': mem.created_at.isoformat() if mem.created_at else None,
        })

    # Bookings with full history
    bookings = []
    for booking in Booking.objects.filter(organization=organization).select_related(
        'customer', 'service',
    ).prefetch_related('status_events', 'request_messages__sender').order_by('created_at'):

        events = []
        for evt in booking.status_events.all():
            events.append({
                'id': evt.id,
                'action': evt.action,
                'old_status': evt.old_status,
                'new_status': evt.new_status,
                'note': evt.note,
                'created_at': evt.created_at.isoformat(),
                'actor_id': evt.actor_id,
            })

        messages = []
        for msg in booking.request_messages.all():
            messages.append({
                'id': msg.id,
                'sender_id': msg.sender_id,
                'sender_name': msg.sender.full_name if msg.sender else None,
                'sender_email': msg.sender.email if msg.sender else None,
                'body': msg.body,
                'created_at': msg.created_at.isoformat(),
            })

        bookings.append({
            'id': booking.id,
            'customer_id': booking.customer_id,
            'customer_name': booking.customer.full_name if booking.customer else None,
            'customer_email': booking.customer.email if booking.customer else None,
            'service_id': booking.service_id,
            'service_name': booking.service.name if booking.service else None,
            'start_at': booking.start_at.isoformat() if booking.start_at else None,
            'end_at': booking.end_at.isoformat() if booking.end_at else None,
            'status': booking.status,
            'service_address': booking.service_address,
            'quote_amount': str(booking.quote_amount) if booking.quote_amount else None,
            'quote_message': booking.quote_message,
            'customer_notes': booking.customer_notes,
            'parent_booking_id': booking.parent_booking_id,
            'created_at': booking.created_at.isoformat(),
            'status_events': events,
            'messages': messages,
        })

    # Service inquiries (custom requests)
    inquiries = []
    for inq in CustomerServiceInquiry.objects.filter(
        organization=organization,
    ).select_related('customer', 'service').prefetch_related('request_messages__sender'):
        inq_messages = []
        for msg in inq.request_messages.all():
            inq_messages.append({
                'id': msg.id,
                'sender_id': msg.sender_id,
                'sender_name': msg.sender.full_name if msg.sender else None,
                'sender_email': msg.sender.email if msg.sender else None,
                'body': msg.body,
                'created_at': msg.created_at.isoformat(),
            })

        inquiries.append({
            'id': inq.id,
            'customer_id': inq.customer_id,
            'customer_name': inq.customer.full_name if inq.customer else None,
            'customer_email': inq.customer.email if inq.customer else None,
            'service_id': inq.service_id,
            'service_label': inq.service_label,
            'message': inq.message,
            'preferred_date': inq.preferred_date.isoformat() if inq.preferred_date else None,
            'service_address': inq.service_address,
            'status': inq.status,
            'created_at': inq.created_at.isoformat(),
            'messages': inq_messages,
        })

    # Invoices
    invoices = []
    for inv in Invoice.objects.filter(
        booking__organization=organization,
    ).select_related('booking', 'booking__customer', 'booking__service'):
        invoices.append({
            'id': inv.id,
            'number': inv.number,
            'booking_id': inv.booking_id,
            'customer_name': inv.booking.customer.full_name if inv.booking.customer else None,
            'customer_email': inv.booking.customer.email if inv.booking.customer else None,
            'service_name': inv.booking.service.name if inv.booking.service else None,
            'description': inv.description,
            'amount': str(inv.amount),
            'subtotal': str(inv.subtotal) if inv.subtotal else None,
            'tax_total': str(inv.tax_total) if inv.tax_total else None,
            'tax_lines': inv.tax_lines,
            'line_items': inv.line_items,
            'currency': inv.currency,
            'status': inv.status,
            'payment_method': inv.payment_method,
            'issued_at': inv.issued_at.isoformat() if inv.issued_at else None,
            'paid_at': inv.paid_at.isoformat() if inv.paid_at else None,
            'qbo_invoice_id': inv.qbo_invoice_id,
        })

    # Job costs
    job_costs = []
    for cost in JobCostLine.objects.filter(
        booking__organization=organization,
    ).select_related('booking', 'booking__customer'):
        job_costs.append({
            'id': cost.id,
            'booking_id': cost.booking_id,
            'customer_name': cost.booking.customer.full_name if cost.booking.customer else None,
            'kind': cost.kind,
            'description': cost.description,
            'quantity': float(cost.quantity) if cost.quantity else None,
            'unit_cost': str(cost.unit_cost) if cost.unit_cost else None,
            'total_cost': str(cost.total_cost) if cost.total_cost else None,
            'created_at': cost.created_at.isoformat() if cost.created_at else None,
        })

    # Service reviews
    reviews = []
    for review in ServiceReview.objects.filter(
        service__organization=organization,
    ).select_related('customer', 'service', 'booking'):
        reviews.append({
            'id': review.id,
            'service_id': review.service_id,
            'service_name': review.service.name,
            'customer_id': review.customer_id,
            'customer_name': review.customer.full_name if review.customer else None,
            'booking_id': review.booking_id,
            'communication': review.communication,
            'price': review.price,
            'punctual': review.punctual,
            'quality': review.quality,
            'comment': review.comment,
            'created_at': review.created_at.isoformat(),
        })

    # Tasks
    tasks = []
    for task in Task.objects.filter(organization=organization).select_related('job'):
        tasks.append({
            'id': task.id,
            'title': task.title,
            'notes': task.notes,
            'job_id': task.job_id,
            'priority': task.priority,
            'due_at': task.due_at.isoformat() if task.due_at else None,
            'is_done': task.is_done,
            'done_at': task.done_at.isoformat() if task.done_at else None,
            'created_at': task.created_at.isoformat(),
        })
    
    return {
        'export_date': export_date.isoformat(),
        'export_version': '1.0',
        'migration_guide': migration_readme_text(),
        'organization': org_data,
        'locations': locations,
        'gallery': gallery,
        'staff': staff,
        'categories': categories,
        'services': services,
        'schedule': {
            'weekly_hours': weekly_hours,
            'unavailable_blocks': unavailable,
            'booked_slots': slots,
        },
        'customers': customers,
        'bookings': bookings,
        'inquiries': inquiries,
        'invoices': invoices,
        'job_costs': job_costs,
        'reviews': reviews,
        'tasks': tasks,
    }


def _cell_value(value: Any) -> Any:
    """Coerce nested values for CSV/Excel cells."""
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False)
    return value


def _flat_row(row: dict[str, Any]) -> dict[str, Any]:
    return {key: _cell_value(value) for key, value in row.items()}


def export_as_json(data: dict[str, Any]) -> bytes:
    """Export data as a single JSON file."""
    return json.dumps(data, indent=2, ensure_ascii=False).encode('utf-8')


def export_as_csv_zip(data: dict[str, Any]) -> bytes:
    """
    Export data as a ZIP file containing multiple CSV files.
    
    One CSV per entity type (organization.csv, customers.csv, bookings.csv, etc.)
    """
    zip_buffer = io.BytesIO()
    
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        zf.writestr('README.txt', migration_readme_text())

        # Organization (single row)
        org_csv = io.StringIO()
        org_writer = csv.DictWriter(org_csv, fieldnames=data['organization'].keys())
        org_writer.writeheader()
        org_writer.writerow(_flat_row(data['organization']))
        zf.writestr('organization.csv', org_csv.getvalue())
        
        # Locations
        if data['locations']:
            loc_csv = io.StringIO()
            loc_writer = csv.DictWriter(loc_csv, fieldnames=data['locations'][0].keys())
            loc_writer.writeheader()
            loc_writer.writerows(data['locations'])
            zf.writestr('locations.csv', loc_csv.getvalue())
        
        # Gallery
        if data['gallery']:
            gallery_csv = io.StringIO()
            gallery_writer = csv.DictWriter(gallery_csv, fieldnames=data['gallery'][0].keys())
            gallery_writer.writeheader()
            gallery_writer.writerows(data['gallery'])
            zf.writestr('gallery.csv', gallery_csv.getvalue())
        
        # Staff
        if data['staff']:
            staff_csv = io.StringIO()
            staff_writer = csv.DictWriter(staff_csv, fieldnames=data['staff'][0].keys())
            staff_writer.writeheader()
            staff_writer.writerows(data['staff'])
            zf.writestr('staff.csv', staff_csv.getvalue())
        
        # Categories
        if data['categories']:
            cat_csv = io.StringIO()
            cat_writer = csv.DictWriter(cat_csv, fieldnames=data['categories'][0].keys())
            cat_writer.writeheader()
            cat_writer.writerows(data['categories'])
            zf.writestr('categories.csv', cat_csv.getvalue())
        
        # Services (flatten gallery into separate rows or skip for CSV simplicity)
        if data['services']:
            # Simplified services without nested gallery
            svc_rows = []
            for svc in data['services']:
                svc_copy = svc.copy()
                svc_copy.pop('gallery', None)
                svc_rows.append(svc_copy)
            
            svc_csv = io.StringIO()
            svc_writer = csv.DictWriter(svc_csv, fieldnames=svc_rows[0].keys())
            svc_writer.writeheader()
            svc_writer.writerows(svc_rows)
            zf.writestr('services.csv', svc_csv.getvalue())
        
        # Schedule hours
        if data['schedule']['weekly_hours']:
            hours_csv = io.StringIO()
            hours_writer = csv.DictWriter(hours_csv, fieldnames=data['schedule']['weekly_hours'][0].keys())
            hours_writer.writeheader()
            hours_writer.writerows(data['schedule']['weekly_hours'])
            zf.writestr('schedule_hours.csv', hours_csv.getvalue())
        
        # Unavailable blocks
        if data['schedule']['unavailable_blocks']:
            unavail_csv = io.StringIO()
            unavail_writer = csv.DictWriter(unavail_csv, fieldnames=data['schedule']['unavailable_blocks'][0].keys())
            unavail_writer.writeheader()
            unavail_writer.writerows(data['schedule']['unavailable_blocks'])
            zf.writestr('unavailable_blocks.csv', unavail_csv.getvalue())
        
        # Customers
        if data['customers']:
            cust_csv = io.StringIO()
            cust_writer = csv.DictWriter(cust_csv, fieldnames=data['customers'][0].keys())
            cust_writer.writeheader()
            cust_writer.writerows(data['customers'])
            zf.writestr('customers.csv', cust_csv.getvalue())
        
        # Bookings (flatten status_events and messages)
        if data['bookings']:
            booking_rows = []
            for bk in data['bookings']:
                bk_copy = bk.copy()
                bk_copy.pop('status_events', None)
                bk_copy.pop('messages', None)
                booking_rows.append(bk_copy)
            
            bk_csv = io.StringIO()
            bk_writer = csv.DictWriter(bk_csv, fieldnames=booking_rows[0].keys())
            bk_writer.writeheader()
            bk_writer.writerows(booking_rows)
            zf.writestr('bookings.csv', bk_csv.getvalue())
            
            # Booking status events
            event_rows = []
            for bk in data['bookings']:
                for evt in bk['status_events']:
                    evt_copy = evt.copy()
                    evt_copy['booking_id'] = bk['id']
                    event_rows.append(evt_copy)
            if event_rows:
                evt_csv = io.StringIO()
                evt_writer = csv.DictWriter(evt_csv, fieldnames=event_rows[0].keys())
                evt_writer.writeheader()
                evt_writer.writerows(event_rows)
                zf.writestr('booking_status_events.csv', evt_csv.getvalue())
        
        # Messages (from bookings + inquiries)
        message_rows = []
        for bk in data['bookings']:
            for msg in bk['messages']:
                msg_copy = msg.copy()
                msg_copy['booking_id'] = bk['id']
                msg_copy['inquiry_id'] = None
                message_rows.append(msg_copy)
        for inq in data['inquiries']:
            for msg in inq['messages']:
                msg_copy = msg.copy()
                msg_copy['booking_id'] = None
                msg_copy['inquiry_id'] = inq['id']
                message_rows.append(msg_copy)
        if message_rows:
            msg_csv = io.StringIO()
            msg_writer = csv.DictWriter(msg_csv, fieldnames=message_rows[0].keys())
            msg_writer.writeheader()
            msg_writer.writerows(message_rows)
            zf.writestr('messages.csv', msg_csv.getvalue())
        
        # Inquiries (flatten messages)
        if data['inquiries']:
            inq_rows = []
            for inq in data['inquiries']:
                inq_copy = inq.copy()
                inq_copy.pop('messages', None)
                inq_rows.append(inq_copy)
            
            inq_csv = io.StringIO()
            inq_writer = csv.DictWriter(inq_csv, fieldnames=inq_rows[0].keys())
            inq_writer.writeheader()
            inq_writer.writerows(inq_rows)
            zf.writestr('inquiries.csv', inq_csv.getvalue())
        
        # Invoices
        if data['invoices']:
            inv_csv = io.StringIO()
            inv_writer = csv.DictWriter(inv_csv, fieldnames=data['invoices'][0].keys())
            inv_writer.writeheader()
            inv_writer.writerows(data['invoices'])
            zf.writestr('invoices.csv', inv_csv.getvalue())
        
        # Job costs
        if data['job_costs']:
            cost_csv = io.StringIO()
            cost_writer = csv.DictWriter(cost_csv, fieldnames=data['job_costs'][0].keys())
            cost_writer.writeheader()
            cost_writer.writerows(data['job_costs'])
            zf.writestr('job_costs.csv', cost_csv.getvalue())
        
        # Reviews
        if data['reviews']:
            rev_csv = io.StringIO()
            rev_writer = csv.DictWriter(rev_csv, fieldnames=data['reviews'][0].keys())
            rev_writer.writeheader()
            rev_writer.writerows(data['reviews'])
            zf.writestr('reviews.csv', rev_csv.getvalue())
        
        # Tasks
        if data['tasks']:
            task_csv = io.StringIO()
            task_writer = csv.DictWriter(task_csv, fieldnames=data['tasks'][0].keys())
            task_writer.writeheader()
            task_writer.writerows(data['tasks'])
            zf.writestr('tasks.csv', task_csv.getvalue())
    
    return zip_buffer.getvalue()


def export_as_excel(data: dict[str, Any]) -> bytes:
    """
    Export data as an Excel workbook with multiple sheets.
    
    One sheet per entity type matching CSV structure.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError("openpyxl is required for Excel export")
    
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # README first so openers see migration guidance immediately
    readme_ws = wb.create_sheet('README', 0)
    readme_ws.column_dimensions['A'].width = 100
    for line in migration_readme_text().splitlines():
        readme_ws.append([line])

    def add_sheet(name: str, rows: list[dict], wb: Workbook):
        if not rows:
            return
        flat_rows = [_flat_row(row) for row in rows]
        ws = wb.create_sheet(name)
        headers = list(flat_rows[0].keys())
        ws.append(headers)

        for row_data in flat_rows:
            ws.append([row_data.get(h) for h in headers])

        # Auto-size columns (approximate)
        for idx, col in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(idx)].width = min(50, max(12, len(col) + 2))

    # Add sheets
    add_sheet('Organization', [data['organization']], wb)
    add_sheet('Locations', data['locations'], wb)
    add_sheet('Gallery', data['gallery'], wb)
    add_sheet('Staff', data['staff'], wb)
    add_sheet('Categories', data['categories'], wb)
    
    # Services (flatten gallery)
    svc_rows = []
    for svc in data['services']:
        svc_copy = svc.copy()
        svc_copy.pop('gallery', None)
        svc_rows.append(svc_copy)
    add_sheet('Services', svc_rows, wb)
    
    add_sheet('Schedule Hours', data['schedule']['weekly_hours'], wb)
    add_sheet('Unavailable Blocks', data['schedule']['unavailable_blocks'], wb)
    add_sheet('Customers', data['customers'], wb)
    
    # Bookings (flatten nested)
    booking_rows = []
    for bk in data['bookings']:
        bk_copy = bk.copy()
        bk_copy.pop('status_events', None)
        bk_copy.pop('messages', None)
        booking_rows.append(bk_copy)
    add_sheet('Bookings', booking_rows, wb)
    
    # Booking status events
    event_rows = []
    for bk in data['bookings']:
        for evt in bk['status_events']:
            evt_copy = evt.copy()
            evt_copy['booking_id'] = bk['id']
            event_rows.append(evt_copy)
    add_sheet('Booking Status Events', event_rows, wb)
    
    # Messages
    message_rows = []
    for bk in data['bookings']:
        for msg in bk['messages']:
            msg_copy = msg.copy()
            msg_copy['booking_id'] = bk['id']
            msg_copy['inquiry_id'] = None
            message_rows.append(msg_copy)
    for inq in data['inquiries']:
        for msg in inq['messages']:
            msg_copy = msg.copy()
            msg_copy['booking_id'] = None
            msg_copy['inquiry_id'] = inq['id']
            message_rows.append(msg_copy)
    add_sheet('Messages', message_rows, wb)
    
    # Inquiries (flatten)
    inq_rows = []
    for inq in data['inquiries']:
        inq_copy = inq.copy()
        inq_copy.pop('messages', None)
        inq_rows.append(inq_copy)
    add_sheet('Inquiries', inq_rows, wb)
    
    add_sheet('Invoices', data['invoices'], wb)
    add_sheet('Job Costs', data['job_costs'], wb)
    add_sheet('Reviews', data['reviews'], wb)
    add_sheet('Tasks', data['tasks'], wb)
    
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    return excel_buffer.getvalue()
